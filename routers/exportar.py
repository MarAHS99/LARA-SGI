"""
Router: /api/cierre/exportar
Genera un Excel del cierre del día con:
  - Sección VENTAS con proveedores dinámicos, unidades y kilos
  - Sección CUENTAS CORRIENTES debajo
Devuelve JSON con filename y data en base64.
"""

import io
import base64
from fastapi import APIRouter, Depends, Query, HTTPException
from sqlalchemy.orm import Session
from sqlalchemy import func
from datetime import date

from database import get_db
from models import Boleta, ProductoBoleta, Compra, ProductoCompra, Proveedor, PagoProveedor, GastoAchurero

def formatMoney_py(n: float) -> str:
    """Formato pesos argentinos para texto en celdas Excel."""
    return f"${n:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")

router = APIRouter(tags=["exportar"])

PRODUCTOS_EXCEL = [
    ('Hígado',        'HIGADO'),
    ('Corazón',       'CORAZON'),
    ('Lengua',        'LENGUA'),
    ('Riñón',         'RIÑON'),
    ('Sesos',         'SESOS'),
    ('Chinchulin',    'CHINCHULIN'),
    ('Tripas rueda',  'TRIPAS'),
    ('Molleja',       'MOLLEJAS'),
    ('Rabo',          'RABO'),
    ('C. de entraña', 'ENTRAÑA'),
    ('Quijada',       'QUIJADAS'),
    ('Mondongo',      'MONDONGO'),
    ('Carne chica',   'CARNE CHI'),
    ('Otros',         'OTROS'),
]

LOC_ORDER   = ['M', 'MDP', 'O', 'B']
LOC_NOMBRES = {'M': 'Miramar', 'MDP': 'Mar del Plata', 'O': 'Otamendi', 'B': 'Balcarce'}


def col_letra(n: int) -> str:
    """Convierte índice 1-based a letra de columna. A=1, B=2..."""
    result = ''
    while n > 0:
        n, rem = divmod(n - 1, 26)
        result = chr(65 + rem) + result
    return result


@router.get("/cierre/exportar")
def exportar_cierre(
    fecha_desde: str = Query(default=None),
    fecha_hasta: str = Query(default=None),
    db: Session = Depends(get_db),
):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="openpyxl no está instalado. Ejecutá: pip install openpyxl"
        )

    hoy = date.today().isoformat()
    if not fecha_desde:
        fecha_desde = hoy
    if not fecha_hasta:
        fecha_hasta = fecha_desde

    if fecha_desde == fecha_hasta:
        label_rango = fecha_desde
    else:
        label_rango = f"{fecha_desde}_al_{fecha_hasta}"

    # ── Ventas del día por producto ───────────────────────────
    rows_ventas = (
        db.query(
            ProductoBoleta.nombre,
            func.sum(ProductoBoleta.cant).label('cant'),
            func.sum(ProductoBoleta.kg).label('kg'),
            func.sum(ProductoBoleta.subtotal).label('subtotal'),
        )
        .join(Boleta, ProductoBoleta.boleta_id == Boleta.id)
        .filter(Boleta.fecha >= fecha_desde, Boleta.fecha <= fecha_hasta)
        .group_by(ProductoBoleta.nombre)
        .all()
    )
    ventas = {
        r.nombre: {
            'cant':     round(float(r.cant or 0)),
            'kg':       round(float(r.kg), 3),
            'subtotal': round(float(r.subtotal), 2),
        }
        for r in rows_ventas
    }

    # ── Compras del día por proveedor y producto ──────────────
    proveedores_db = db.query(Proveedor).order_by(Proveedor.nombre).all()
    prov_nombres   = [p.nombre for p in proveedores_db]

    compras_dia  = {p: {} for p in prov_nombres}
    compras_cant = {p: {} for p in prov_nombres}
    rows_compras = (
        db.query(
            Compra.proveedor,
            ProductoCompra.nombre,
            func.sum(ProductoCompra.cant).label('cant'),
            func.sum(ProductoCompra.kg).label('kg'),
        )
        .join(ProductoCompra, ProductoCompra.compra_id == Compra.id)
        .filter(Compra.fecha >= fecha_desde, Compra.fecha <= fecha_hasta)
        .group_by(Compra.proveedor, ProductoCompra.nombre)
        .all()
    )
    for r in rows_compras:
        if r.proveedor not in compras_dia:
            compras_dia[r.proveedor]  = {}
            compras_cant[r.proveedor] = {}
        compras_dia[r.proveedor][r.nombre]  = round(float(r.kg), 3)
        compras_cant[r.proveedor][r.nombre] = round(float(r.cant or 0))

    # Proveedores que tuvieron compras hoy; si ninguno, mostrar todos
    provs_activos = [p for p in prov_nombres if compras_dia.get(p)]
    if not provs_activos:
        provs_activos = prov_nombres

    # ── Cuentas corrientes ────────────────────────────────────
    from routers.cuentas import _calcular_saldos
    saldos = _calcular_saldos(db)

    # ── Helpers de estilo ─────────────────────────────────────
    def fill(hex_color):
        return PatternFill('solid', start_color=hex_color, end_color=hex_color)

    def fnt(bold=False, color='000000', size=10):
        return Font(name='Arial', bold=bold, color=color, size=size)

    def brd():
        s = Side(style='thin', color='AAAAAA')
        return Border(left=s, right=s, top=s, bottom=s)

    def aln(h='left'):
        return Alignment(horizontal=h, vertical='center')

    def cell(coord, value=None, bold=False, color='000000', size=10,
             bg=None, h='left', border=True, num_fmt=None):
        c = ws[coord]
        if value is not None:
            c.value = value
        c.font      = fnt(bold=bold, color=color, size=size)
        c.alignment = aln(h)
        if bg:
            c.fill = fill(bg)
        if border:
            c.border = brd()
        if num_fmt:
            c.number_format = num_fmt
        return c

    BG_H   = 'D9D9D9'
    BG_LOC = '2E75B6'
    BG_TOT = 'E2EFDA'
    BG_GEN = 'BDD7EE'
    BG_W   = 'FFFFFF'
    RED    = 'C00000'
    GREEN  = '375623'

    # ── Índices de columnas dinámicos ─────────────────────────
    # A=precio prom, B=pedido, C=articulos
    # D,E = prov1  F,G = prov2  etc.
    # Luego: J=unid total compra, K=kg total compra
    # L=unid ventas, M=articulos ventas, N=kilos ventas, O=subtotal ventas
    COL_A = 1; COL_B = 2; COL_C = 3
    COL_PROV_START = 4                          # D
    n_provs = len(provs_activos)
    COL_J   = COL_PROV_START + n_provs * 2     # después de todos los proveedores
    COL_K   = COL_J + 1
    COL_L   = COL_K + 1
    COL_M   = COL_L + 1
    COL_N   = COL_M + 1
    COL_O   = COL_N + 1

    CJ = col_letra(COL_J); CK = col_letra(COL_K)
    CL = col_letra(COL_L); CM = col_letra(COL_M)
    CN = col_letra(COL_N); CO = col_letra(COL_O)

    # ── Workbook ──────────────────────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = f'CIERRE {label_rango}'

    ws.column_dimensions['A'].width = 16.2
    ws.column_dimensions['B'].width = 22.58   # 210px
    ws.column_dimensions['C'].width = 23.22   # 216px
    ws.column_dimensions['D'].width = 13.33   # 124px
    ws.column_dimensions['E'].width = 13.33   # 124px
    ws.column_dimensions['F'].width = 16.24   # 151px
    for i in range(n_provs):
        ws.column_dimensions[col_letra(COL_PROV_START + i*2)].width     = 12.5
        ws.column_dimensions[col_letra(COL_PROV_START + i*2 + 1)].width = 12.5
    ws.column_dimensions[CJ].width = 13
    ws.column_dimensions[CK].width = 13
    ws.column_dimensions[CL].width = 13
    ws.column_dimensions[CM].width = 15
    ws.column_dimensions[CN].width = 12
    ws.column_dimensions[CO].width = 16

    # ── Fila 1 ────────────────────────────────────────────────
    if fecha_desde == fecha_hasta:
        cell('B1', 'FECHA', bold=True, bg=BG_H, h='center')
        cell('C1', fecha_desde, bold=True, size=11, h='center')
    else:
        cell('B1', 'PERÍODO', bold=True, bg=BG_H, h='center')
        cell('C1', f'{fecha_desde}  →  {fecha_hasta}', bold=True, size=10, h='center')
    cell(f'{CL}1', 'VENTA TOTAL', bold=True, bg=BG_H, h='center')

    # ── Fila 2 — headers ──────────────────────────────────────
    cell('A2', 'PRECIO',    bold=True, bg=BG_H, h='center')
    cell('B2', 'PEDIDO',    bold=True, bg=BG_H, h='center')
    cell('C2', 'ARTICULOS', bold=True, bg=BG_H, h='center')

    for i, prov in enumerate(provs_activos):
        cu = col_letra(COL_PROV_START + i*2)
        cell(f'{cu}2', prov.upper(), bold=True, bg=BG_H, h='center')

    cell(f'{CJ}2', 'UNID.',     bold=True, bg=BG_H, h='center')
    cell(f'{CK}2', 'KGS',       bold=True, bg=BG_H, h='center')
    cell(f'{CL}2', 'Unid.',     bold=True, bg=BG_H, h='center')
    cell(f'{CM}2', 'ARTICULOS', bold=True, bg=BG_H, h='center')
    cell(f'{CN}2', 'KILOS',     bold=True, bg=BG_H, h='center')
    cell(f'{CO}2', 'SUB. TOTAL',bold=True, bg=BG_H, h='center')

    # ── Fila 3 — subheaders ───────────────────────────────────
    cell('A3', 'PROMEDIO', bold=True, bg=BG_H, h='center')
    for i in range(n_provs):
        cu = col_letra(COL_PROV_START + i*2)
        ck = col_letra(COL_PROV_START + i*2 + 1)
        cell(f'{cu}3', 'Unid.', bold=True, bg=BG_H, h='center')
        cell(f'{ck}3', 'KGS',   bold=True, bg=BG_H, h='center')
    cell(f'{CM}3', 'ARTICULOS', bold=True, bg=BG_H, h='center')
    cell(f'{CN}3', 'KILOS',     bold=True, bg=BG_H, h='center')
    cell(f'{CO}3', 'SUB. TOTAL',bold=True, bg=BG_H, h='center')

    # ── Filas 4+ — productos ──────────────────────────────────
    fila = 4
    for nombre_lara, nombre_excel in PRODUCTOS_EXCEL:
        d      = ventas.get(nombre_lara, {'cant': 0, 'kg': 0.0, 'subtotal': 0.0})
        cant   = d['cant']
        kg     = d['kg']
        subtot = d['subtotal']

        cell(f'C{fila}',  nombre_excel, bg=BG_W)
        cell(f'{CM}{fila}', nombre_excel, bg=BG_W)

        # Kg por proveedor
        for i, prov in enumerate(provs_activos):
            ck = col_letra(COL_PROV_START + i*2 + 1)
            kg_prov = compras_dia.get(prov, {}).get(nombre_lara)
            c = ws[f'{ck}{fila}']
            c.value = kg_prov
            c.font = fnt(); c.fill = fill(BG_W)
            c.alignment = aln('right'); c.border = brd()
            c.number_format = '#,##0.000'

        # Total unidades compradas → CJ
        total_cant_compra = sum(compras_cant.get(prov, {}).get(nombre_lara) or 0 for prov in provs_activos)
        c = ws[f'{CJ}{fila}']
        c.value = int(total_cant_compra) if total_cant_compra > 0 else None
        c.font = fnt(); c.fill = fill(BG_W)
        c.alignment = aln('right'); c.border = brd()

        # Total kg comprados → CK
        total_kg_compra = sum(compras_dia.get(prov, {}).get(nombre_lara) or 0 for prov in provs_activos)
        c = ws[f'{CK}{fila}']
        c.value = round(total_kg_compra, 3) if total_kg_compra > 0 else None
        c.font = fnt(); c.fill = fill(BG_W)
        c.alignment = aln('right'); c.border = brd()
        c.number_format = '#,##0.000'

        # Unidades vendidas
        c = ws[f'{CL}{fila}']
        c.value = int(cant) if cant > 0 else None
        c.font = fnt(); c.fill = fill(BG_W)
        c.alignment = aln('right'); c.border = brd()

        # Kilos vendidos
        c = ws[f'{CN}{fila}']
        c.value = kg if kg > 0 else None
        c.font = fnt(); c.fill = fill(BG_W)
        c.alignment = aln('right'); c.border = brd()
        c.number_format = '#,##0.000'

        # Subtotal ventas
        c = ws[f'{CO}{fila}']
        c.value = subtot if subtot > 0 else None
        c.font = fnt(); c.fill = fill(BG_W)
        c.alignment = aln('right'); c.border = brd()
        c.number_format = '$#,##0.00'

        # Precio promedio
        c = ws[f'A{fila}']
        c.value = f'=IFERROR({CO}{fila}/{CN}{fila},"")'
        c.font = fnt(); c.fill = fill(BG_W)
        c.alignment = aln('right'); c.border = brd()
        c.number_format = '$#,##0.00'

        fila += 1

    fila_fin_prod = fila - 1

    # ── Fila total ventas ─────────────────────────────────────
    cell(f'{CL}{fila}', 'TOTAL:', bold=True, bg=BG_GEN, h='right', size=13)

    c = ws[f'{CN}{fila}']
    c.value = f'=SUM({CN}4:{CN}{fila_fin_prod})'
    c.font = fnt(bold=True, size=13); c.fill = fill(BG_GEN)
    c.alignment = aln('right'); c.border = brd()
    c.number_format = '#,##0.000'

    c = ws[f'{CO}{fila}']
    c.value = f'=SUM({CO}4:{CO}{fila_fin_prod})'
    c.font = fnt(bold=True, size=13); c.fill = fill(BG_GEN)
    c.alignment = aln('right'); c.border = brd()
    c.number_format = '$#,##0.00'

    fila += 3

    # ════════════════════════════════════════
    # SECCIÓN CUENTAS CORRIENTES
    # ════════════════════════════════════════
    ws[f'A{fila}'] = 'CUENTAS CORRIENTES'
    ws[f'A{fila}'].font = fnt(bold=True, size=12)
    ws[f'A{fila}'].fill = fill(BG_H)
    ws[f'A{fila}'].alignment = aln('left')
    ws.merge_cells(f'A{fila}:F{fila}')
    fila += 1

    for col, txt in [('A','CLIENTE'),('B','FECHA'),('C','OBSERVACION'),
                     ('D','DEBE'),('E','HABER'),('F','SALDO')]:
        cell(f'{col}{fila}', txt, bold=True, bg=BG_H, h='center')
    fila += 1

    total_general = 0.0

    for loc in LOC_ORDER:
        if loc not in saldos or not saldos[loc]:
            continue

        ws[f'A{fila}'] = LOC_NOMBRES.get(loc, loc)
        ws[f'A{fila}'].font = fnt(bold=True, color='FFFFFF', size=11)
        ws[f'A{fila}'].fill = fill(BG_LOC)
        ws[f'A{fila}'].alignment = aln('left')
        ws.merge_cells(f'A{fila}:F{fila}')
        fila += 1

        clientes_loc = sorted(saldos[loc].items(), key=lambda x: x[1]['saldo_actual'], reverse=True)
        fila_ini = fila

        for cli, datos in clientes_loc:
            sa = datos['saldo_actual']
            total_general += sa

            cell(f'A{fila}', cli,      bg=BG_W)
            cell(f'B{fila}', label_rango, bg=BG_W, h='center')
            cell(f'C{fila}', 'ACHURAS',bg=BG_W)

            c = ws[f'D{fila}']
            c.value = datos['saldo_boletas'] if datos['saldo_boletas'] > 0 else None
            c.font = fnt(color=RED); c.fill = fill(BG_W)
            c.alignment = aln('right'); c.border = brd()
            c.number_format = '$#,##0.00'

            c = ws[f'E{fila}']
            c.value = datos['total_pagado'] if datos['total_pagado'] > 0 else None
            c.font = fnt(color=GREEN); c.fill = fill(BG_W)
            c.alignment = aln('right'); c.border = brd()
            c.number_format = '$#,##0.00'

            c = ws[f'F{fila}']
            c.value = sa
            c.font = fnt(bold=True, color=RED if sa > 0 else GREEN)
            c.fill = fill(BG_W); c.alignment = aln('right'); c.border = brd()
            c.number_format = '$#,##0.00'
            fila += 1

        fila_fin  = fila - 1
        total_loc = sum(d['saldo_actual'] for _, d in clientes_loc)

        ws[f'A{fila}'] = f'Total {LOC_NOMBRES.get(loc, loc)}'
        ws[f'A{fila}'].font      = fnt(bold=True)
        ws[f'A{fila}'].fill      = fill(BG_TOT)
        ws[f'A{fila}'].alignment = aln('right')
        ws.merge_cells(f'A{fila}:E{fila}')

        c = ws[f'F{fila}']
        c.value = f'=SUM(F{fila_ini}:F{fila_fin})'
        c.font = fnt(bold=True, color=RED if total_loc > 0 else GREEN, size=11)
        c.fill = fill(BG_TOT); c.alignment = aln('right'); c.border = brd()
        c.number_format = '$#,##0.00'
        fila += 2

    ws[f'A{fila}'] = 'TOTAL GENERAL'
    ws[f'A{fila}'].font      = fnt(bold=True, size=12)
    ws[f'A{fila}'].fill      = fill(BG_GEN)
    ws[f'A{fila}'].alignment = aln('right')
    ws.merge_cells(f'A{fila}:E{fila}')

    c = ws[f'F{fila}']
    c.value = total_general
    c.font = fnt(bold=True, color=RED if total_general > 0 else GREEN, size=13)
    c.fill = fill(BG_GEN); c.alignment = aln('right')
    c.number_format = '$#,##0.00'

    # ════════════════════════════════════════
    # SECCIÓN GASTOS DE ACHUREROS (columnas H-L)
    # ════════════════════════════════════════
    gastos_rows = (
        db.query(GastoAchurero)
        .filter(GastoAchurero.fecha >= fecha_desde, GastoAchurero.fecha <= fecha_hasta)
        .order_by(GastoAchurero.fecha)
        .all()
    )

    if gastos_rows:
        fila_ga = 22  # siempre empieza en fila 22, al lado de la grilla de productos

        # Header sección
        ws[f'H{fila_ga}'] = 'GASTOS ACHUREROS'
        ws[f'H{fila_ga}'].font      = fnt(bold=True, size=12)
        ws[f'H{fila_ga}'].fill      = fill(BG_H)
        ws[f'H{fila_ga}'].alignment = aln('left')
        ws.merge_cells(f'H{fila_ga}:L{fila_ga}')
        fila_ga += 1

        # Subheaders
        for col, txt in [('H','FECHA'),('I','ACHURERO'),('J','LOCALIDAD'),('K','MONTO'),('L','NOTA')]:
            cell(f'{col}{fila_ga}', txt, bold=True, bg=BG_H, h='center')
        fila_ga += 1

        fila_ini_ga = fila_ga
        total_gastos = 0.0

        for g in gastos_rows:
            total_gastos += g.monto
            cell(f'H{fila_ga}', g.fecha,                              bg=BG_W, h='center')
            cell(f'I{fila_ga}', g.achurero,                           bg=BG_W)
            cell(f'J{fila_ga}', LOC_NOMBRES.get(g.locacion, g.locacion or '—'), bg=BG_W, h='center')

            c = ws[f'K{fila_ga}']
            c.value = round(g.monto, 2)
            c.font = fnt(color=RED); c.fill = fill(BG_W)
            c.alignment = aln('right'); c.border = brd()
            c.number_format = '$#,##0.00'

            cell(f'L{fila_ga}', g.nota or '—', bg=BG_W)
            fila_ga += 1

        fila_fin_ga = fila_ga - 1

        # Total gastos
        ws[f'H{fila_ga}'] = 'TOTAL GASTOS'
        ws[f'H{fila_ga}'].font      = fnt(bold=True)
        ws[f'H{fila_ga}'].fill      = fill(BG_TOT)
        ws[f'H{fila_ga}'].alignment = aln('right')
        ws.merge_cells(f'H{fila_ga}:J{fila_ga}')

        c = ws[f'K{fila_ga}']
        c.value = f'=SUM(K{fila_ini_ga}:K{fila_fin_ga})'
        c.font = fnt(bold=True, color=RED, size=11)
        c.fill = fill(BG_TOT); c.alignment = aln('right'); c.border = brd()
        c.number_format = '$#,##0.00'
        fila_ga += 2

        # Anchos columnas H-L
        ws.column_dimensions['H'].width = 13
        ws.column_dimensions['I'].width = 18
        ws.column_dimensions['J'].width = 16
        ws.column_dimensions['K'].width = 14
        ws.column_dimensions['L'].width = 22

        # ── CC PROVEEDORES (columnas H-L, debajo de gastos) ──
        fila_ga += 1  # espacio visual
        ws[f'H{fila_ga}'] = 'CUENTAS CORRIENTES PROVEEDORES'
        ws[f'H{fila_ga}'].font      = fnt(bold=True, size=12)
        ws[f'H{fila_ga}'].fill      = fill(BG_H)
        ws[f'H{fila_ga}'].alignment = aln('left')
        ws.merge_cells(f'H{fila_ga}:L{fila_ga}')
        fila_ga += 1

        for col_p, txt_p in [('H','PROVEEDOR'),('I','PERÍODO'),('J','COMPRAS'),('K','PAGADO'),('L','SALDO')]:
            cell(f'{col_p}{fila_ga}', txt_p, bold=True, bg=BG_H, h='center')
        fila_ga += 1

        proveedores_cc = db.query(Proveedor).order_by(Proveedor.nombre).all()
        total_cc_prov  = 0.0
        fila_ini_prov  = fila_ga

        for prov in proveedores_cc:
            comp_prov = float(db.query(func.coalesce(func.sum(Compra.total), 0))
                .filter(Compra.proveedor == prov.nombre).scalar() or 0)
            pago_prov = float(db.query(func.coalesce(func.sum(PagoProveedor.monto), 0))
                .filter(PagoProveedor.proveedor == prov.nombre).scalar() or 0)
            saldo_prov = round(comp_prov - pago_prov, 2)
            if comp_prov == 0 and pago_prov == 0:
                continue
            total_cc_prov += saldo_prov

            cell(f'H{fila_ga}', prov.nombre,   bg=BG_W)
            cell(f'I{fila_ga}', label_rango,    bg=BG_W, h='center')

            for col_v, val_v, clr in [
                ('J', comp_prov,  RED),
                ('K', pago_prov,  GREEN),
                ('L', saldo_prov, RED if saldo_prov > 0 else GREEN),
            ]:
                cv = ws[f'{col_v}{fila_ga}']
                cv.value = round(val_v, 2) if val_v > 0 else None
                cv.font = fnt(bold=(col_v=='L'), color=clr if val_v > 0 else '888888')
                cv.fill = fill(BG_W); cv.alignment = aln('right'); cv.border = brd()
                cv.number_format = '$#,##0.00'
            fila_ga += 1

        fila_fin_prov = fila_ga - 1

        # Total CC proveedores
        ws[f'H{fila_ga}'] = 'TOTAL ADEUDADO'
        ws[f'H{fila_ga}'].font      = fnt(bold=True)
        ws[f'H{fila_ga}'].fill      = fill(BG_TOT)
        ws[f'H{fila_ga}'].alignment = aln('right')
        ws.merge_cells(f'H{fila_ga}:K{fila_ga}')
        cv = ws[f'L{fila_ga}']
        cv.value = round(total_cc_prov, 2)
        cv.font = fnt(bold=True, color=RED if total_cc_prov > 0 else GREEN, size=11)
        cv.fill = fill(BG_TOT); cv.alignment = aln('right'); cv.border = brd()
        cv.number_format = '$#,##0.00'

        # ── MARGEN NETO TERMINAL (columnas H-L, bien destacado) ──
        fila_ga += 3
        total_ventas_excel = float(
            db.query(func.coalesce(func.sum(Boleta.total), 0))
            .filter(Boleta.fecha >= fecha_desde, Boleta.fecha <= fecha_hasta)
            .scalar() or 0
        )
        total_compras_excel = float(
            db.query(func.coalesce(func.sum(Compra.total), 0))
            .filter(Compra.fecha >= fecha_desde, Compra.fecha <= fecha_hasta)
            .scalar() or 0
        )
        margen_neto = round(total_ventas_excel - total_compras_excel - total_gastos, 2)
        color_neto  = GREEN if margen_neto >= 0 else RED
        BG_NETO     = 'D9EAD3' if margen_neto >= 0 else 'F4CCCC'  # verde/rojo suave

        ws[f'H{fila_ga}'] = '💰 MARGEN NETO DEL PERÍODO'
        ws[f'H{fila_ga}'].font      = fnt(bold=True, size=14)
        ws[f'H{fila_ga}'].fill      = fill(BG_NETO)
        ws[f'H{fila_ga}'].alignment = aln('center')
        ws.merge_cells(f'H{fila_ga}:K{fila_ga}')
        ws.row_dimensions[fila_ga].height = 28

        cv = ws[f'L{fila_ga}']
        cv.value = margen_neto
        cv.font = fnt(bold=True, color=color_neto, size=16)
        cv.fill = fill(BG_NETO); cv.alignment = aln('center'); cv.border = brd()
        cv.number_format = '$#,##0.00'
        ws.row_dimensions[fila_ga].height = 28

        fila_ga += 1
        ws[f'H{fila_ga}'] = f'Ventas {formatMoney_py(total_ventas_excel)}  −  Compras {formatMoney_py(total_compras_excel)}  −  Gastos achureros {formatMoney_py(total_gastos)}'
        ws[f'H{fila_ga}'].font      = fnt(size=9, color='666666')
        ws[f'H{fila_ga}'].fill      = fill(BG_NETO)
        ws[f'H{fila_ga}'].alignment = aln('center')
        ws.merge_cells(f'H{fila_ga}:L{fila_ga}')

    # ── Serializar a base64 ───────────────────────────────────
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    b64 = base64.b64encode(buffer.read()).decode('utf-8')

    return {
        "filename": f"LARA_Cierre_{label_rango}.xlsx",
        "data":     b64,
    }